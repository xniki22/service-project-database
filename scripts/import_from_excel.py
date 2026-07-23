from datetime import date, datetime
from pathlib import Path
import sqlite3

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
DATABASE_PATH = BASE_DIR / "data" / "service_projects.db"
EXCEL_PATH = BASE_DIR / "data" / "service_projects_input.xlsx"

REQUIRED_COLUMNS = {
    "Visiting Program",
    "Program Year",
    "Project Code",
    "Project Name",
    "Project Date",
    "Organization",
    "Number of Participants",
    "Number of Hours",
    "Notes",
    "Supporting Documents",
}


def clean_text(value) -> str:
    """Convert an Excel value to trimmed text."""

    if value is None:
        return ""

    return str(value).strip()


def format_date(value) -> str:
    """Convert an Excel date to YYYY-MM-DD format."""

    if value is None or value == "":
        return ""

    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")

    text_value = clean_text(value)

    try:
        return datetime.strptime(text_value, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError as error:
        raise ValueError(
            f"Invalid date '{text_value}'. Use YYYY-MM-DD."
        ) from error


def get_or_create_program(
    cursor: sqlite3.Cursor,
    program_name: str,
    program_year: int,
) -> int:
    """Return an existing program ID or create a new program."""

    cursor.execute(
        """
        SELECT program_id
        FROM programs
        WHERE program_name = ?
          AND program_year = ?
        """,
        (program_name, program_year),
    )

    row = cursor.fetchone()

    if row:
        return row[0]

    cursor.execute(
        """
        INSERT INTO programs (
            program_name,
            program_year,
            notes
        )
        VALUES (?, ?, ?)
        """,
        (program_name, program_year, None),
    )

    return cursor.lastrowid


def get_or_create_organization(
    cursor: sqlite3.Cursor,
    organization_name: str,
) -> int | None:
    """Return an existing organization ID or create a new organization."""

    if not organization_name:
        return None

    cursor.execute(
        """
        SELECT organization_id
        FROM organizations
        WHERE organization_name = ?
        """,
        (organization_name,),
    )

    row = cursor.fetchone()

    if row:
        return row[0]

    cursor.execute(
        """
        INSERT INTO organizations (
            organization_name,
            location
        )
        VALUES (?, ?)
        """,
        (organization_name, None),
    )

    return cursor.lastrowid


def import_excel_data() -> None:
    """Insert or update service projects using the Excel input file."""

    if not DATABASE_PATH.exists():
        print(f"Database not found: {DATABASE_PATH}")
        print("Run scripts/create_database.py first.")
        return

    if not EXCEL_PATH.exists():
        print(f"Excel input file not found: {EXCEL_PATH}")
        return

    workbook = load_workbook(EXCEL_PATH, data_only=True)
    worksheet = workbook.active

    headers = {
        clean_text(cell.value): column_number
        for column_number, cell in enumerate(worksheet[1], start=1)
        if clean_text(cell.value)
    }

    missing_columns = REQUIRED_COLUMNS - set(headers)

    if missing_columns:
        print("The spreadsheet is missing required columns:")
        for column in sorted(missing_columns):
            print(f"  - {column}")
        return

    inserted_count = 0
    updated_count = 0
    skipped_count = 0

    try:
        with sqlite3.connect(DATABASE_PATH) as connection:
            connection.execute("PRAGMA foreign_keys = ON;")
            cursor = connection.cursor()

            for row_number in range(2, worksheet.max_row + 1):
                def value(column_name: str):
                    return worksheet.cell(
                        row=row_number,
                        column=headers[column_name],
                    ).value

                program_name = clean_text(value("Visiting Program"))
                project_code = clean_text(value("Project Code"))
                project_name = clean_text(value("Project Name"))

                # Ignore completely empty spreadsheet rows.
                if not program_name and not project_code and not project_name:
                    continue

                try:
                    if not program_name:
                        raise ValueError("Visiting Program is required.")

                    if not project_code:
                        raise ValueError("Project Code is required.")

                    if not project_name:
                        raise ValueError("Project Name is required.")

                    program_year_value = value("Program Year")

                    if program_year_value in (None, ""):
                        raise ValueError("Program Year is required.")

                    program_year = int(program_year_value)

                    project_date = format_date(value("Project Date"))

                    participant_value = value("Number of Participants")
                    participant_count = (
                        int(participant_value)
                        if participant_value not in (None, "")
                        else None
                    )

                    hours_value = value("Number of Hours")
                    hours_per_participant = (
                        float(hours_value)
                        if hours_value not in (None, "")
                        else None
                    )

                    organization_name = clean_text(value("Organization"))
                    notes = clean_text(value("Notes"))
                    document_url = clean_text(
                        value("Supporting Documents")
                    )

                    program_id = get_or_create_program(
                        cursor,
                        program_name,
                        program_year,
                    )

                    organization_id = get_or_create_organization(
                        cursor,
                        organization_name,
                    )

                    cursor.execute(
                        """
                        SELECT project_id
                        FROM service_projects
                        WHERE project_code = ?
                        """,
                        (project_code,),
                    )

                    existing_project = cursor.fetchone()

                    if existing_project:
                        project_id = existing_project[0]

                        cursor.execute(
                            """
                            UPDATE service_projects
                            SET
                                project_name = ?,
                                program_id = ?,
                                organization_id = ?,
                                project_date = ?,
                                participant_count = ?,
                                hours_per_participant = ?,
                                notes = ?,
                                sync_status = 'Not Synced'
                            WHERE project_id = ?
                            """,
                            (
                                project_name,
                                program_id,
                                organization_id,
                                project_date or None,
                                participant_count,
                                hours_per_participant,
                                notes or None,
                                project_id,
                            ),
                        )

                        updated_count += 1
                    else:
                        cursor.execute(
                            """
                            INSERT INTO service_projects (
                                project_code,
                                project_name,
                                program_id,
                                organization_id,
                                project_date,
                                participant_count,
                                hours_per_participant,
                                notes,
                                sync_status
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                project_code,
                                project_name,
                                program_id,
                                organization_id,
                                project_date or None,
                                participant_count,
                                hours_per_participant,
                                notes or None,
                                "Not Synced",
                            ),
                        )

                        project_id = cursor.lastrowid
                        inserted_count += 1

                    if document_url:
                        cursor.execute(
                            """
                            SELECT document_id
                            FROM supporting_documents
                            WHERE project_id = ?
                              AND box_url = ?
                            """,
                            (project_id, document_url),
                        )

                        if cursor.fetchone() is None:
                            cursor.execute(
                                """
                                INSERT INTO supporting_documents (
                                    project_id,
                                    document_name,
                                    document_type,
                                    box_url,
                                    description
                                )
                                VALUES (?, ?, ?, ?, ?)
                                """,
                                (
                                    project_id,
                                    "Supporting Documents",
                                    "Box Link",
                                    document_url,
                                    None,
                                ),
                            )

                except (TypeError, ValueError) as error:
                    skipped_count += 1
                    print(f"Row {row_number} skipped: {error}")

            connection.commit()

        print("Excel import completed.")
        print(f"Inserted: {inserted_count}")
        print(f"Updated: {updated_count}")
        print(f"Skipped: {skipped_count}")
        print(f"Database updated: {DATABASE_PATH}")

    except sqlite3.Error as error:
        print(f"Database error: {error}")

    finally:
        workbook.close()


if __name__ == "__main__":
    import_excel_data()